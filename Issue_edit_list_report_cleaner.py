# ==========================================================
# Issue Edit List Report Cleaner
# Author : Muhammad Waqas
# Purpose: Convert messy data of Issue Edit List report .CSV into Cleaned Excel Report
# ==========================================================

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==========================================================
# File Paths
# ==========================================================

INPUT_FILE = "/sample_data/Issue_edit_list.csv"
OUTPUT_FILE = "/sample_data/Cleaned_issue_edit_list.xlsx"

# ==========================================================
# Regular Expressions
# ==========================================================

ACCOUNT_PATTERN = r"^203-\d+"
ITEM_PATTERN = r"^\d{2}-"

# ==========================================================
# Read CSV File
# ==========================================================

def read_csv(File_Path):

    print("=" * 60)
    print("Reading CSV File...")
    print("=" * 60)

    df = pd.read_csv(
        File_Path,
        header=None,
        dtype=str,
        keep_default_na=False
    )

    print("Rows Read :", len(df))
    print()

    return df

# ==========================================================
# Safe Column Reader
# ==========================================================

def get_value(row, column_index):

    if column_index >= len(row):
        return ""

    return str(row[column_index]).strip()
# ==========================================================
# Extract Transactions
# ==========================================================

def extract_transactions(df):

    records = []

    account_code = ""
    account_name = ""

    print("=" * 60)
    print("Extracting Transactions...")
    print("=" * 60)

    # Read file row by row
    for row_index in range(len(df)):

        row = df.iloc[row_index]

        # Read required columns safely
        col0 = get_value(row, 0)
        col2 = get_value(row, 2)
        col3 = get_value(row, 3)

        # --------------------------------------------------
        # Account Row
        # --------------------------------------------------

        if re.match(ACCOUNT_PATTERN, col0):

            account_code = col0
            account_name = col2

            continue

        # --------------------------------------------------
        # Transaction Row
        # --------------------------------------------------

        if col3 and re.match(ITEM_PATTERN, col3):

            records.append({

                "Account Code": account_code,

                "Account Name": account_name,

                "Transaction Date": col0,

                "Creation Date": get_value(row, 1),

                "SIR No": col2,

                "Item Code": col3,

                "Item Description": get_value(row, 4),

                "UOM": get_value(row, 5),

                "Quantity": get_value(row, 6),

                "Amount": get_value(row, 7)

            })

    print(f"Transactions Found : {len(records)}")
    print()

    return pd.DataFrame(records)
    # ==========================================================
# Clean Text Columns
# ==========================================================

def clean_text_columns(df):

    print("=" * 60)
    print("Cleaning Text Columns...")
    print("=" * 60)

    for column in df.columns:

        df[column] = (
            df[column]
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    return df


# ==========================================================
# Clean Numeric Columns
# ==========================================================

def clean_numeric_columns(df):

    print("Cleaning Numeric Columns...")

    # Remove commas
    df["Quantity"] = (
        df["Quantity"]
        .astype(str)
        .str.replace(",", "", regex=False)
    )

    df["Amount"] = (
        df["Amount"]
        .astype(str)
        .str.replace(",", "", regex=False)
    )

    # Convert to numbers
    df["Quantity"] = pd.to_numeric(
        df["Quantity"],
        errors="coerce"
    ).fillna(0)

    df["Amount"] = pd.to_numeric(
        df["Amount"],
        errors="coerce"
    ).fillna(0)

    return df


# ==========================================================
# Calculate Rate
# ==========================================================

def calculate_rate(df):

    print("Calculating Rate...")

    df["Rate"] = df.apply(

        lambda row:
        round(row["Amount"] / row["Quantity"], 2)
        if row["Quantity"] != 0 else 0,

        axis=1
    )

    return df


# ==========================================================
# Sort Records
# ==========================================================

def sort_records(df):

    print("Sorting Records...")

    df = df.sort_values(
        by="SIR No"
    ).reset_index(drop=True)

    return df

# ==========================================================
# Save Excel File
# ==========================================================

def save_excel(df, output_file):

    print("=" * 60)
    print("Saving Excel File...")
    print("=" * 60)

    df.to_excel(output_file, index=False)

    print()
    print("SUCCESS!")
    print(f"Output File : {output_file}")
    print(f"Total Records : {len(df)}")

# ==========================================================
# Format Excel File
# ==========================================================
def format_excel(output_file):

    print("=" * 60)
    print("Formatting Excel File...")
    print("=" * 60)

    # Open Workbook
    wb = load_workbook(output_file)
    ws = wb.active
    ws.title = "Report"

    # Create Excel Table
    last_row = ws.max_row
    last_col = ws.max_column

    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

    tab = Table(
        displayName="ReportTable",
        ref=table_ref
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Freeze Header
    ws.freeze_panes = "A2"

    # Auto Fit Columns
    for column_cells in ws.columns:

        length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:

            try:
                if cell.value is not None:
                    length = max(length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column].width = min(length + 3, 50)

    # Header Formatting
    for cell in ws[1]:

        cell.font = Font(bold=True)

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Save
    wb.save(output_file)



# ==========================================================
# Main Function
# ==========================================================

def main():

    # Read CSV
    df = read_csv(INPUT_FILE)

    # Extract Transactions
    final_df = extract_transactions(df)

    # Check if transactions exist
    if final_df.empty:

        print("No transaction records found.")
        return

    # Clean Text
    final_df = clean_text_columns(final_df)

    # Clean Numbers
    final_df = clean_numeric_columns(final_df)

    # Calculate Rate
    final_df = calculate_rate(final_df)

    # Sort Records
    final_df = sort_records(final_df)

    # Save Output
    save_excel(final_df, OUTPUT_FILE)

    #Formatting Output
    format_excel(OUTPUT_FILE)


    print()
    print("=" * 60)
    print("Issue Edit List Cleaning Completed Successfully")
    print("=" * 60)


# ==========================================================
# Run Program
# ==========================================================

if __name__ == "__main__":
    main()