# ==========================================================
# Store Section Wise Report Cleaner
# Author : Muhammad Waqas
# Purpose: Convert messy data of Store Section Wise Report .CSV into Cleaned Excel File
# ==========================================================
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==========================================================
# READ CSV FILE
# ==========================================================

def read_csv_file(file_path):

    print("=" * 60)
    print(f"Reading CSV File: {file_path}...")
    print("=" * 60)

    try:
        df = pd.read_csv(
            file_path,
            header=None,
            dtype=str,
            keep_default_na=False
        )
        return df
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return None
    except Exception as e:
        print(f"An error occurred while reading the CSV file: {e}")
        return None

# ==========================================================
# Cleaning White Spaces
# ==========================================================
def clean_whitespace(df):
    print("Clean Whitespaces...")
    print("=" * 60)
    for col in df.columns:
      df[col] = (
        df[col]
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return df

# ==========================================================
# Extraction Records
# ==========================================================
def extract_records(df):
    print("Record Extracted...")
    print("=" * 60)

    # VARIABLES
    records = []
    headers = []
    current_category = ""
    rows = len(df)
    cols = len(df.columns)
    i = 0

    # MAIN LOOP
    while i < rows:
        first_cell = str(df.iloc[i, 0]).strip()

        # FIND ITEM ROW
        if first_cell.upper() == "ITEM":

            # Previous row contains category
            if i > 0:
                current_category = str(df.iloc[i - 1, 0]).strip()

            # Merge Header Row + Code Row
            if i + 1 < rows:

                header1 = df.iloc[i].tolist()
                header2 = df.iloc[i + 1].tolist()

                headers = []

                for h1, h2 in zip(header1, header2):

                    h1 = str(h1).strip()
                    h2 = str(h2).strip()

                    if h1 and h2:
                        headers.append(f"{h1} {h2}")

                    elif h1:
                        headers.append(h1)

                    elif h2:
                        headers.append(h2)

                    else:
                        headers.append("")

            i += 2
            continue


        # FIND ITEM CODE
        if re.match(r"^\d{5}-", first_cell):
            row = df.iloc[i].tolist()

            # Prefix 01-
            row[0] = "01-" + first_cell

            # Merge Multi-line Description
            description = str(row[1]).strip()
            j = i + 1
            while j < rows:
                next_first = str(df.iloc[j, 0]).strip()

                # Stop on next item code
                if re.match(r"^\d{5}-", next_first):
                    break

                # Stop on next ITEM section
                if next_first.upper() == "ITEM":
                    break

                # Merge description text
                extra = str(df.iloc[j, 1]).strip()

                if extra:
                    description += " " + extra

                j += 1

            description = re.sub(r"\s+", " ", description).strip()

            row[1] = description

            # Add Category at first column
            final_row = [current_category] + row

            records.append(final_row)

            i = j
            continue

        i += 1

    # CREATE HEADERS
    if headers:

        headers[0] = "Item Code"
        headers[1] = "Item Description"

        final_headers = ["Category"] + headers

    else:

        final_headers = ["Category"]

        for n in range(cols):
            final_headers.append(f"Column_{n+1}")

    return records, final_headers, cols

# ==========================================================
# CREATE DATAFRAME
# ==========================================================
def create_dataframe(records, final_headers):
    print("Create Dataframe...")
    print("=" * 60)
    final_df = pd.DataFrame(records, columns=final_headers)
    return final_df

# ==========================================================
# FINAL CLEANING
# ==========================================================
def convert_numeric_columns(final_df, numeric_cols_to_convert):

    print("Set Data Types...")
    print("=" * 60)
    for col in final_df.columns:

        final_df[col] = (
            final_df[col]
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    # Remove completely blank rows
    final_df = final_df.replace("", pd.NA).dropna(how="all").fillna("")

    # Convert Quantity, Value Columns to Numeric
    for col in numeric_cols_to_convert:
        if col in final_df.columns:
            final_df[col] = (
                final_df[col]
                .astype(str)
                .str.replace(",", "", regex=False)   # Remove commas
                .str.strip()
            )

            final_df[col] = pd.to_numeric(final_df[col], errors="coerce").fillna(0)
    return final_df
# ==========================================================
# SAVE TO EXCEL
# ==========================================================
def save_file(df_to_save, file_path):
    print("Save To Excel File...")
    try:
        df_to_save.to_excel(file_path, index=False)
        print("Cleaned File Save at:",file_path)
        return file_path
    except IOError as e:
        print(f"Error: Could not save the Excel file to '{file_path}'. {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while saving the Excel file: {e}")
        return None

# ==========================================================
# Format Excel File
# ==========================================================
def format_excel(output_file):
    print("=" * 60)
    print("Formatting Excel File...")
    print("=" * 60)

    try:
        # Open Workbook
        wb = load_workbook(output_file)
        ws = wb.active
        ws.title = "Report"

        # Create Excel Table
        last_row = ws.max_row
        last_col = ws.max_column

        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

        tbl = Table(
            displayName="ReportTbl",
            ref=table_ref
        )

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tbl.tableStyleInfo = style
        ws.add_table(tbl)

        # Freeze Header
        ws.freeze_panes = "A2"

        # Auto Fit Columns
        for column_cells in ws.columns:

            length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:

                try:
                    if cell.value is not None:
                        length = max(length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[column].width = min(length + 3, 50)

        # Header Formatting
        for cell in ws[1]:

            cell.font = Font(bold=True)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # Save
        wb.save(output_file)
        print(f"Excel file formatted and saved to {output_file}")
    except FileNotFoundError:
        print(f"Error: The output file '{output_file}' was not found for formatting.")
    except Exception as e:
        print(f"An error occurred during Excel formatting: {e}")

# ==========================================================
# Main Function
# ==========================================================

def main(input_csv_path="/sample_data/Store_Section_Wise_Capitalized.csv",
         output_excel_path="/sample_data/Cleaned_Store_Section_Wise_Capitalized.xlsx",
         numeric_cols_to_convert=[
            "Consumption Quantity",
            "Consumption Value",
            "Capitalized Quantity",
            "Capitalized Value",
            "Total Quantity",
            "Total Value"
        ]):

    # Read CSV
    df = read_csv_file(input_csv_path)
    if df is None:
        return # Exit if CSV reading failed

    #Clean White_Spaces
    df = clean_whitespace(df)

    # Extract Records
    records, headers, cols = extract_records(df)

    # Creating DataFrame
    final_df = create_dataframe(records, headers)

    # Set Data Types
    final_df = convert_numeric_columns(final_df, numeric_cols_to_convert)

    # Save File Into Excel
    saved_file_path = save_file(final_df, output_excel_path)
    if saved_file_path is None:
        return # Exit if saving to Excel failed

    #Formatting Output
    format_excel(output_excel_path)

    print("Summary",)
    print("="*60)
    print("Sum of Consumption Quantities:",final_df["Consumption Quantity"].sum())
    print("Sum of Consumption Values:",final_df["Consumption Value"].sum())
    print("Sum of Capitalized Quantities:",final_df["Capitalized Quantity"].sum())
    print("Sum of Capitalized Values:",final_df["Capitalized Value"].sum())
    print("Sum of Total Quantities:",final_df["Total Quantity"].sum())
    print("Sum of Total Values:",final_df["Total Value"].sum())




# ==========================================================
# Run Program
# ==========================================================

if __name__ == "__main__":
    main()